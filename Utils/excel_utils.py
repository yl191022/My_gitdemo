# excel 封装
import openpyxl


class Excel:
    @classmethod
    def excel_write(cls, fine_name, msg):
        try:
            w = openpyxl.Workbook()
            dyb = w.active
            for x in msg:
                dyb.append(x)
            w.save(fine_name + '.xlsx')
            w.close()
        except Exception as er:
            print('写入失败:', er)

    @classmethod
    def excel_read(cls, fine_name, start_row=1, start_column=1):
        try:
            r = openpyxl.load_workbook(fine_name + '.xlsx')
            sheet = r.active
            all_data = []
            for row in range(start_row, sheet.max_row+1):
                tmp = []
                for column in range(start_column, sheet.max_column+1):
                    tmp.append(sheet.cell(row, column).value)
                all_data.append(tmp)
            r.close()
            return all_data
        except Exception as er:
            print('读取失败:', er)


# da_list = [[random.randint(1, 10000) for x in range(6)] for y in range(10)]
# Excel.excel_write('TEST_exce', da_list)
# print(Excel.excel_read('TEST_exce'))
